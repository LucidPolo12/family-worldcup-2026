#!/usr/bin/env python3
"""Regenerate data.json for the Family World Cup portal from FamilyWorldCup2026.xlsx.

data.json is the LIVE data source the portal fetches on load. After any change to
the spreadsheet (results or picks), run:  python3 regenerate_data.py
Then publish (commit/push or upload data.json). The portal picks it up automatically.
"""
import openpyxl, json, os, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "FamilyWorldCup2026.xlsx")
OUT  = os.path.join(HERE, "data.json")

# Player -> (Pred-Home col, Pred-Away col) in the Match Predictions sheet
PCOLS = [("Ewan","I","J"),("Nana","L","M"),("Emrys","O","P"),("Nonno","R","S"),
         ("Nonna","U","V"),("Boompa","X","Y"),("Xavier","AA","AB"),
         ("Autumn","AD","AE"),("River","AG","AH")]
CHAMPS = {"Ewan":"Spain","Nana":"France","Nonno":"Brazil","Boompa":"France"}
PHOTOS = {"Nonno":"nonno-bobblehead.jpg","Emrys":"emrys-bobblehead.jpg",
          "Nana":"nana-bobblehead.jpg","Nonna":"nonna-bobblehead.jpg",
          "Boompa":"boompa-bobblehead.jpg","Autumn":"autumn-bobblehead.jpg",
          "River":"river-bobblehead.jpg","Xavier":"xavier-bobblehead.jpg",
          "Ewan":"ewan-bobblehead.jpg"}

# ---- Scoring tables -----------------------------------------------------------
# GROUP STAGE (unchanged): correct outcome = 3, exact score = 8 (exact REPLACES
# the 3, it does not stack). Judged by score sign (win / draw / loss).
GROUP_PTS = (3, 8)   # (correct_outcome, exact_score_total)

# KNOCKOUT STAGE (new model): two INDEPENDENT, ADDITIVE components per match —
#   (1) advance_pts  if you named the team that actually advances (incl. via
#       penalties); for a non-tied predicted score the advancing side is the
#       higher score, for a tied predicted score it is the player's explicit
#       penalty pick.
#   (2) exact_bonus  if your predicted final score (after all playing time,
#       penalties excluded) equals the actual final score.
# Round of 32 is 5 + 7 = 12. Later rounds keep the previously-announced totals,
# split into advance + exact-bonus (advance == old "correct result" value).
#   (advance_pts, exact_bonus)  ->  max per match = advance_pts + exact_bonus
KO_PTS = {
    "Round of 32":  (5,  7),    # max 12
    "Round of 16":  (8,  10),   # max 18
    "Quarterfinal": (12, 13),   # max 25
    "Semifinal":    (18, 17),   # max 35
    "Third Place":  (12, 13),   # max 25
    "Final":        (25, 25),   # max 50
}

def match_stage(n):
    """Return the stage name for match number n (1-104)."""
    if n <= 72:  return "Group Stage"
    if n <= 88:  return "Round of 32"
    if n <= 96:  return "Round of 16"
    if n <= 100: return "Quarterfinal"
    if n <= 102: return "Semifinal"
    if n == 103: return "Third Place"
    return "Final"

def adv_side(h, a, manual=None):
    """Which side advances given a FINAL playing-time score.
    'H' = home/Team-A side, 'A' = away/Team-B side, None = unknown.
    Decisive score -> the higher side. Tied score (penalty shootout) -> the
    'manual' override (H/A/home/away), since the score alone can't tell us."""
    if h is None or a is None:
        return None
    if h > a: return "H"
    if a > h: return "A"
    m = ("" if manual is None else str(manual).strip().upper())
    if m in ("H", "HOME"): return "H"
    if m in ("A", "AWAY"): return "A"
    return None   # tied & penalty winner not recorded yet

def compute_ranks(standings):
    """Rank purely by points (ties share a rank), matching the front-end logic."""
    s = sorted(standings, key=lambda x: (-x["matchPts"], -x["exact"], -x["correct"], x["name"]))
    ranks, rank = {}, 1
    for i, x in enumerate(s):
        if i > 0 and x["matchPts"] < s[i-1]["matchPts"]:
            rank = i + 1
        ranks[x["name"]] = rank
    return ranks

def pred_adv_side(ph, pa, padv=None):
    """Predicted advancing side from a player's pick.
    Non-tied predicted score -> higher side (auto). Tied predicted score ->
    the player's explicit penalty pick ('H'/'A'); None if they didn't choose."""
    if ph is None or pa is None:
        return None
    if ph > pa: return "H"
    if pa > ph: return "A"
    return padv if padv in ("H", "A") else None

def pts(ph, pa, ah, aa, n=1, padv=None, aadv=None):
    """Return (points, is_exact). Points is None if result not yet known.

    Group stage: unchanged (3 correct outcome / 8 exact, exact replaces).
    Knockout: advance_pts (correct advancing team, penalties included) PLUS
    exact_bonus (exact final playing-time score) — the two stack independently.
    `padv` = player's penalty pick side for a tied predicted score;
    `aadv` = the side that actually advanced (from adv_side())."""
    if None in (ph, pa, ah, aa):
        return None, False
    stage = match_stage(n)
    if stage == "Group Stage":
        correct_pts, exact_pts = GROUP_PTS
        if ph == ah and pa == aa: return exact_pts, True
        sign = lambda x, y: (x > y) - (x < y)
        return (correct_pts if sign(ph, pa) == sign(ah, aa) else 0), False
    # ---- knockout ----
    advance_pts, exact_bonus = KO_PTS[stage]
    total = 0
    p_side = pred_adv_side(ph, pa, padv)
    if p_side is not None and aadv is not None and p_side == aadv:
        total += advance_pts
    is_exact = (ph == ah and pa == aa)
    if is_exact:
        total += exact_bonus
    return total, is_exact

def score_golden_boot(gb, top_scorers):
    """Score Golden Boot picks against the official top-3 scorer list.

    top_scorers: list of player names in rank order (index 0 = Golden Boot winner).
    Scoring: pick1 exact (rank 1) = 40pts; pick1 in top-3 = 10pts;
             pick2/3 hits rank-1 = 20pts; pick2/3 in top-3 = 10pts.
    Returns 0 if top_scorers not yet populated.
    """
    if not top_scorers or not gb:
        return 0
    total = 0
    for pick_pos, pick_key in enumerate(["pick1", "pick2", "pick3"], 1):
        player = gb.get(pick_key, "").strip()
        if not player:
            continue
        try:
            actual_rank = top_scorers.index(player) + 1   # 1-based
        except ValueError:
            continue   # player not in top scorers list
        if actual_rank > 3:
            continue   # only top 3 count
        if pick_pos == 1 and actual_rank == 1:
            total += 40   # nailed the Golden Boot winner exactly
        elif pick_pos == 1:
            total += 10   # pick1 got a top-3 scorer (not #1)
        elif actual_rank == 1:
            total += 20   # pick2/3 named the actual Golden Boot winner
        else:
            total += 10   # pick2/3 got a top-3 scorer
    return total

def main():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Match Predictions"]

    # Results come from results.json (owned & auto-filled by the GitHub Action) merged
    # with any scores typed directly into the spreadsheet's G/H columns. results.json wins.
    rstore = {}
    rpath = os.path.join(HERE, "results.json")
    if os.path.exists(rpath):
        try:
            rstore = {int(k): v for k, v in json.load(open(rpath, encoding="utf-8")).items()}
        except Exception:
            rstore = {}

    # Family-submitted picks (from the Google Sheet via fetch_picks.py) override the
    # spreadsheet's pick cells per player+match.
    pjson = {}
    ppath = os.path.join(HERE, "picks.json")
    if os.path.exists(ppath):
        try:
            pjson = json.load(open(ppath, encoding="utf-8")) or {}
        except Exception:
            pjson = {}

    # A knockout match that ends level is decided on penalties; the score alone
    # can't say who advanced, so the admin records the advancing SIDE by hand in
    # the new "Advancing (H/A)" column (AJ) of the Match Predictions sheet — or as
    # an optional 3rd element in a results.json entry ([h, a, "H"|"A"]).
    results, adv_manual = {}, {}
    for r in range(4, 108):
        n = r - 3
        g, h = ws["G"+str(r)].value, ws["H"+str(r)].value
        rv = rstore.get(n)
        if rv is not None:
            results[n] = [rv[0], rv[1]]
            if len(rv) > 2 and rv[2]:
                adv_manual[n] = rv[2]
        elif g is not None and h is not None:
            results[n] = [g, h]
        aj = ws["AJ"+str(r)].value          # manual penalty-winner override
        if aj not in (None, ""):
            adv_manual[n] = aj

    # "generated" = date of the first match with no result yet (the next slate to pick)
    generated = None
    for r in range(4, 108):
        if (r - 3) not in results:
            d = ws["B"+str(r)].value
            generated = d.strftime("%Y-%m-%d") if d else None
            break

    mres = {n: (results[n][0], results[n][1]) if n in results else (None, None)
            for n in range(1, 105)}

    # Actual advancing side per match: auto for decisive scores, manual for ties.
    advancing = {}
    for n in range(73, 105):              # knockout matches only
        ah, aa = mres[n]
        side = adv_side(ah, aa, adv_manual.get(n))
        if side:
            advancing[n] = side

    # Load existing data.json early: need topScorers before the standings loop,
    # and probs + rank snapshots for the output.
    old = {}
    if os.path.exists(OUT):
        try:
            with open(OUT, encoding="utf-8") as f:
                old = json.load(f) or {}
        except Exception:
            old = {}
    probs = old.get("probs", {}) or {}
    top_scorers = old.get("topScorers") or []

    standings = []
    for name, hc, ac in PCOLS:
        picks, mp, ex, cor = [], 0, 0, 0
        pj = pjson.get(name, {})
        for r in range(4, 108):
            n = r - 3
            ov = pj.get(str(n))
            padv = None
            if ov:
                ph, pa = ov[0], ov[1]
                if len(ov) > 2 and ov[2] in ("H", "A"):
                    padv = ov[2]          # explicit penalty pick on a tied score
            else:
                ph, pa = ws[hc+str(r)].value, ws[ac+str(r)].value
            if ph is None and pa is None: continue
            ah, aa = mres[n]
            aadv = advancing.get(n)
            p, is_exact = pts(ph, pa, ah, aa, n, padv=padv, aadv=aadv)
            pk = {"n": n, "ph": ph, "pa": pa, "pts": p}
            # store the predicted advancing side for knockout picks (for display)
            if n >= 73:
                ps = pred_adv_side(ph, pa, padv)
                if ps: pk["adv"] = ps
            picks.append(pk)
            if p is not None:
                mp += p
                if is_exact: ex += 1
                if p > 0: cor += 1

        # Champion & Podium picks, auto-synced from Google Sheet via fetch_picks.py.
        podium = pj.get("podium") or {}
        champ = podium.get("champ") or CHAMPS.get(name)
        gb = pj.get("goldenBoot") or {}
        gb_pts = score_golden_boot(gb, top_scorers)
        standings.append({"name": name, "matchPts": mp, "exact": ex, "correct": cor,
                          "champ": champ, "podium": podium, "goldenBoot": gb, "gbPts": gb_pts,
                          "photo": PHOTOS.get(name), "picks": picks})

    # Day-over-day movement: prevRanks is yesterday's final ranks; ranksToday
    # is the current run. Multiple runs the same day keep the same baseline.
    today_str = datetime.date.today().isoformat()
    cur_ranks = compute_ranks(standings)
    prev_ranks = old.get("prevRanks")
    old_today = old.get("ranksToday")
    if old_today and old_today.get("date") != today_str:
        prev_ranks = old_today

    # stagePts is what the front-end reads. For knockout rounds "correct" is the
    # advance value and "exact" is the per-match max (advance + exact-bonus), so
    # the existing badge/legend logic keeps working with the new additive model.
    stage_pts_out = {"Group Stage": {"correct": GROUP_PTS[0], "exact": GROUP_PTS[1]}}
    for stg, (adv_p, ex_b) in KO_PTS.items():
        stage_pts_out[stg] = {"correct": adv_p, "exact": adv_p + ex_b, "advance": adv_p, "bonus": ex_b}

    data = {
        "generated": generated,
        "standings": standings,
        "results": results,
        "advancing": advancing,
        "photos": PHOTOS,
        "probs": probs,
        "prevRanks": prev_ranks,
        "ranksToday": {"date": today_str, "ranks": cur_ranks},
        "stagePts": stage_pts_out,
        "topScorers": top_scorers,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    print(f"Wrote {OUT}: {len(standings)} players, {len(results)} results, "
          f"{len(probs)} probs kept, generated={generated}")

if __name__ == "__main__":
    main()
